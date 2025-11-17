
from rest_framework import viewsets, status
from rest_framework.permissions import IsAdminUser
from rest_framework.decorators import api_view, permission_classes,action
from rest_framework.response import Response
from rest_framework.views import APIView
from datetime import datetime, date
from django.utils import timezone
from django.middleware.csrf import get_token
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils.decorators import method_decorator
from core.models import TaxRate, Currency
from django.http import HttpResponse
import csv
from django.db.models import Sum, F, Q
from io import BytesIO
from decimal import Decimal
from .models import JournalEntry, JournalLine, BankAccount, InvoiceApproval
from segment.models import XX_Segment
from ar.models import ARInvoice, ARPayment
from ap.models import APInvoice, APPayment
from .serializers import (JournalLineSerializer, JournalLineDetailSerializer, JournalEntrySerializer,ARInvoiceSerializer, ARPaymentSerializer,APInvoiceSerializer, APPaymentSerializer,JournalEntryReadSerializer, BankAccountSerializer,SeedVATRequestSerializer, CorpTaxAccrualRequestSerializer)
from segment.serializers import AccountSerializer
from .services import (
    post_entry,
    gl_post_from_ar_balanced, gl_post_from_ap_balanced,
    post_ar_payment, post_ap_payment,
    reverse_journal, seed_vat_presets, accrue_corporate_tax,
    q2, _with_org_filter, ar_totals, ap_totals
)
from .models import CorporateTaxFiling
from .services import (
    resolve_tax_rate_for_date,
    accrue_corporate_tax_with_filing, reverse_corporate_tax_filing, file_corporate_tax,
    # reuse accrue_corporate_tax if you need raw calc without filing
)
try:
    from openpyxl import Workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


from .services import build_trial_balance, build_ar_aging, build_ap_aging


from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiExample, inline_serializer
from rest_framework import serializers as drf_serializers

@extend_schema(
    parameters=[],
    description="Trial Balance report. Supports ?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD and format choices.",
    responses={200: None}
)
class TrialBalanceReport(APIView):
    def get(self, request):
        df = request.GET.get("date_from")
        dt = request.GET.get("date_to")
        fmt = (request.GET.get("format") or "json").lower()
        file_type = request.GET.get("file_type", "").lower()

        # Force list to avoid generator exhaustion
        rows = list(build_trial_balance(df, dt))

        if fmt == "csv" or file_type == "csv":
            # CSV uses built-in csv module, no external library needed
            resp = HttpResponse(content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="trial_balance.csv"'
            resp.write('\ufeff')  # Excel-friendly BOM
            w = csv.writer(resp)
            w.writerow(["Account Code", "Account Name", "Debit", "Credit"])
            for r in rows:
                w.writerow([
                    r.get("code", ""),
                    r.get("name", ""),
                    f'{r.get("debit", 0):.2f}',
                    f'{r.get("credit", 0):.2f}'
                ])
            return resp

        if fmt in ("xlsx", "excel", "xlsm") or file_type == "xlsx":
            if not OPENPYXL_OK:
               return HttpResponse("openpyxl not installed", status=400)
            wb = Workbook()
            ws = wb.active
            ws.title = "Trial Balance"
            ws.append(["Account Code", "Account Name", "Debit", "Credit"])
            for r in rows:
                ws.append([
                    r.get("code", ""),
                    r.get("name", ""),
                    float(r.get("debit", 0)),
                    float(r.get("credit", 0))
                ])

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            resp = HttpResponse(
                bio.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = 'attachment; filename="trial_balance.xlsx"'
            return resp

        # Default JSON response with download links
        base_url = request.build_absolute_uri(request.path)
        query_params = []
        if df:
            query_params.append(f"date_from={df}")
        if dt:
            query_params.append(f"date_to={dt}")
        
        query_string = "&".join(query_params)
        
        response_data = {
            "data": rows,
            "download_links": {
                "xlsx": f"{base_url}?file_type=xlsx" + (f"&{query_string}" if query_string else ""),
                "csv": f"{base_url}?file_type=csv" + (f"&{query_string}" if query_string else "")
            }
        }
        
        return Response(response_data)


@extend_schema(
    parameters=[],
    description="AR Aging report. Supports ?as_of=YYYY-MM-DD, and bucket sizes b1,b2,b3.",
    responses={200: None}
)
class ARAgingReport(APIView):
    def get(self, request):
        as_of_str = request.GET.get("as_of")
        if as_of_str:
            try:
                as_of = datetime.strptime(as_of_str, "%Y-%m-%d").date()
            except ValueError:
                as_of = date.today()
        else:
            as_of = date.today()

        b1 = int(request.GET.get("b1", 30))
        b2 = int(request.GET.get("b2", 30))
        b3 = int(request.GET.get("b3", 30))
        fmt = (request.GET.get("format") or "json").lower()
        file_type = request.GET.get("file_type", "").lower()

        data = build_ar_aging(as_of, b1, b2, b3)
        
        if fmt == "csv" or file_type == "csv":
            resp = HttpResponse(content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="ar_aging.csv"'
            resp.write('\ufeff')  # Excel-friendly BOM
            w = csv.writer(resp)
            w.writerow(["Invoice ID","Number","Customer","Date","Due Date","Days Overdue","Bucket","Balance"])
            for r in data["invoices"]:
                w.writerow([r["invoice_id"], r["number"], r["customer"], r["date"], r["due_date"], r["days_overdue"], r["bucket"], f'{r["balance"]:.2f}'])
            # add summary rows
            w.writerow([]); w.writerow(["Summary"])
            for k in data["buckets"] + [">=TOTAL_MARKER"]:
                if k == ">=TOTAL_MARKER":
                    w.writerow(["TOTAL", "", "", "", "", "", "", f'{data["summary"]["TOTAL"]:.2f}'])
                else:
                    w.writerow([k, "", "", "", "", "", "", f'{data["summary"][k]:.2f}'])
            return resp

        if fmt in ("xlsx", "excel") or file_type == "xlsx":
            if not OPENPYXL_OK:
               return HttpResponse("openpyxl not installed", status=400)
            wb = Workbook(); ws = wb.active; ws.title = "AR Aging"
            ws.append(["Invoice ID","Number","Customer","Date","Due Date","Days Overdue","Bucket","Balance"])
            for r in data["invoices"]:
                ws.append([r["invoice_id"], r["number"], r["customer"], r["date"], r["due_date"], r["days_overdue"], r["bucket"], r["balance"]])
            ws.append([]); ws.append(["Summary"])
            for k in data["buckets"]:
                ws.append([k, "", "", "", "", "", "", data["summary"][k]])
            ws.append(["TOTAL","","","","","","", data["summary"]["TOTAL"]])
            bio = BytesIO(); wb.save(bio); bio.seek(0)
            resp = HttpResponse(bio.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = 'attachment; filename="ar_aging.xlsx"'
            return resp

        # Default JSON response with download links
        base_url = request.build_absolute_uri(request.path)
        query_params = []
        if as_of_str:
            query_params.append(f"as_of={as_of_str}")
        if request.GET.get("b1"):
            query_params.append(f"b1={b1}")
        if request.GET.get("b2"):
            query_params.append(f"b2={b2}")
        if request.GET.get("b3"):
            query_params.append(f"b3={b3}")
        
        query_string = "&".join(query_params)
        
        response_data = {
            "data": data,
            "download_links": {
                "xlsx": f"{base_url}?file_type=xlsx" + (f"&{query_string}" if query_string else ""),
                "csv": f"{base_url}?file_type=csv" + (f"&{query_string}" if query_string else "")
            }
        }
        
        return Response(response_data)

@extend_schema(
    parameters=[],
    description="AP Aging report. Supports ?as_of=YYYY-MM-DD, and bucket sizes b1,b2,b3.",
    responses={200: None}
)
class APAgingReport(APIView):
    def get(self, request):
        as_of_str = request.GET.get("as_of")
        if as_of_str:
            as_of_str = as_of_str.strip()  # ✅ remove stray newline
            try:
                as_of = datetime.strptime(as_of_str, "%Y-%m-%d").date()
            except ValueError:
                as_of = date.today()
        else:
            as_of = date.today()

        b1 = int(request.GET.get("b1", 30))
        b2 = int(request.GET.get("b2", 30))
        b3 = int(request.GET.get("b3", 30))
        fmt = (request.GET.get("format") or "json").lower()
        file_type = request.GET.get("file_type", "").lower()

        data = build_ap_aging(as_of, b1, b2, b3)

        if fmt == "csv" or file_type == "csv":
            resp = HttpResponse(content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="ap_aging.csv"'
            resp.write('\ufeff')  # Excel-friendly BOM
            w = csv.writer(resp)
            w.writerow(["Invoice ID","Number","Supplier","Date","Due Date","Days Overdue","Bucket","Balance"])
            for r in data["invoices"]:
                w.writerow([r["id"], r["number"], r["supplier"], r["date"], r["due_date"], r["days_overdue"], r["bucket"], r["balance"]])
            # add summary rows
            w.writerow([]); w.writerow(["Summary"])
            for k in data["buckets"] + [">=TOTAL_MARKER"]:
                if k == ">=TOTAL_MARKER":
                    w.writerow(["TOTAL", "", "", "", "", "", "", f'{data["summary"]["TOTAL"]:.2f}'])
                else:
                    w.writerow([k, "", "", "", "", "", "", f'{data["summary"][k]:.2f}'])
            return resp

        if fmt in ("xlsx", "excel") or file_type == "xlsx":
            if not OPENPYXL_OK:
               return HttpResponse("openpyxl not installed", status=400)
            wb = Workbook(); ws = wb.active; ws.title = "AP Aging"
            ws.append(["Invoice ID","Number","Supplier","Date","Due Date","Days Overdue","Bucket","Balance"])
            for r in data["invoices"]:
                ws.append([r["invoice_id"], r["number"], r["supplier"], r["date"], r["due_date"], r["days_overdue"], r["bucket"], r["balance"]])
            ws.append([]); ws.append(["Summary"])
            for k in data["buckets"]:
                ws.append([k, "", "", "", "", "", "", data["summary"][k]])
            ws.append(["TOTAL","","","","","","", data["summary"]["TOTAL"]])
            bio = BytesIO(); wb.save(bio); bio.seek(0)
            resp = HttpResponse(bio.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = 'attachment; filename="ap_aging.xlsx"'
            return resp

        # Default JSON response with download links
        base_url = request.build_absolute_uri(request.path)
        query_params = []
        if as_of_str:
            query_params.append(f"as_of={as_of_str}")
        if request.GET.get("b1"):
            query_params.append(f"b1={b1}")
        if request.GET.get("b2"):
            query_params.append(f"b2={b2}")
        if request.GET.get("b3"):
            query_params.append(f"b3={b3}")
        
        query_string = "&".join(query_params)
        
        response_data = {
            "data": data,
            "download_links": {
                "xlsx": f"{base_url}?file_type=xlsx" + (f"&{query_string}" if query_string else ""),
                "csv": f"{base_url}?file_type=csv" + (f"&{query_string}" if query_string else "")
            }
        }
        
        return Response(response_data)
class BankAccountViewSet(viewsets.ModelViewSet):
    serializer_class = BankAccountSerializer
    queryset = BankAccount.objects.all()


@method_decorator(ensure_csrf_cookie, name='dispatch')
class GetCSRFToken(APIView):
    """
    Get CSRF token for subsequent requests.
    Call this endpoint first to get the CSRF cookie.
    """
    permission_classes = []  # Allow anyone to get CSRF token
    
    def get(self, request):
        # The @ensure_csrf_cookie decorator will set the cookie
        return Response({'detail': 'CSRF cookie set'})


class CurrencyViewSet(viewsets.ModelViewSet):
    from .serializers import CurrencySerializer
    serializer_class = CurrencySerializer
    queryset = Currency.objects.all()
    filterset_fields = ["code", "is_base"]


# AccountViewSet moved to segment/api.py


class JournalEntryViewSet(viewsets.ModelViewSet):
    serializer_class = JournalEntrySerializer; queryset = JournalEntry.objects.all()
    
    @action(detail=True, methods=["post"], url_path="post")
    def post_gl(self, request, pk=None):
        entry = self.get_object(); post_entry(entry); return Response({"status": "posted"})
    
    @action(detail=True, methods=["post"], url_path="reverse")
    def reverse_entry(self, request, pk=None):
        """Reverse a posted journal entry"""
        entry = self.get_object()
        if not entry.posted:
            return Response({"error": "Only posted entries can be reversed"}, status=400)
        reversed_entry = reverse_journal(entry)
        serializer = self.get_serializer(reversed_entry)
        return Response({"status": "reversed", "reversed_entry": serializer.data})

    @extend_schema(
    parameters=[
        # Beware: drf-spectacular infers types from serializers; use explicit params for simple query strings
        OpenApiExample("CSV export", value={"file_type": "csv"}, parameter_only=True),
        OpenApiExample("XLSX export", value={"file_type": "xlsx"}, parameter_only=True),
    ],
    responses={200: None},
    description="Export journal entries to CSV or XLSX. Query params: file_type=csv|xlsx, date_from, date_to, posted_only=true|false"
)
    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """Export journal entries to CSV or XLSX"""
        file_type = request.GET.get("file_type", "csv").lower()
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        posted_only = request.GET.get("posted_only", "false").lower() == "true"

        # Build queryset with filters
        queryset = JournalEntry.objects.all()
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        if posted_only:
            queryset = queryset.filter(posted=True)
        
        queryset = queryset.order_by('date', 'id').prefetch_related('lines__account')

        if file_type == "csv":
            resp = HttpResponse(content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="journals.csv"'
            resp.write('\ufeff')  # BOM for Excel
            w = csv.writer(resp)
            w.writerow(["Journal ID", "Date", "Currency", "Memo", "Posted", "Account Code", "Account Name", "Debit", "Credit"])
            
            for entry in queryset:
                for line in entry.lines.all():
                    w.writerow([
                        entry.id,
                        entry.date,
                        entry.currency.code if entry.currency else "",
                        entry.memo,
                        "Yes" if entry.posted else "No",
                        line.account.code,
                        line.account.name,
                        f"{line.debit:.2f}",
                        f"{line.credit:.2f}"
                    ])
            return resp

        elif file_type == "xlsx":
            if not OPENPYXL_OK:
                return HttpResponse("openpyxl not installed", status=400)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Journal Entries"
            ws.append(["Journal ID", "Date", "Currency", "Memo", "Posted", "Account Code", "Account Name", "Debit", "Credit"])
            
            for entry in queryset:
                for line in entry.lines.all():
                    ws.append([
                        entry.id,
                        entry.date,
                        entry.currency.code if entry.currency else "",
                        entry.memo,
                        "Yes" if entry.posted else "No",
                        line.account.code,
                        line.account.name,
                        float(line.debit),
                        float(line.credit)
                    ])
            
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            resp = HttpResponse(bio.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = 'attachment; filename="journals.xlsx"'
            return resp
        
        else:
            return Response({"error": "Invalid file_type. Use 'csv' or 'xlsx'."}, status=400)

class JournalLineViewSet(viewsets.ReadOnlyModelViewSet):
    """Read-only viewset for journal lines with filtering capabilities"""
    serializer_class = JournalLineDetailSerializer
    queryset = JournalLine.objects.select_related('entry', 'account', 'entry__currency').all()
    filterset_fields = ['account', 'entry']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by account code
        account_code = self.request.query_params.get('account_code')
        if account_code:
            queryset = queryset.filter(account__code__icontains=account_code)
        
        # Filter by account name
        account_name = self.request.query_params.get('account_name')
        if account_name:
            queryset = queryset.filter(account__name__icontains=account_name)
        
        # Filter by date range
        date_from = self.request.query_params.get('date_from')
        if date_from:
            queryset = queryset.filter(entry__date__gte=date_from)
        
        date_to = self.request.query_params.get('date_to')
        if date_to:
            queryset = queryset.filter(entry__date__lte=date_to)
        
        # Filter by posted status
        posted = self.request.query_params.get('posted')
        if posted is not None:
            queryset = queryset.filter(entry__posted=posted.lower() == 'true')
        
        # Filter by journal entry ID
        entry_id = self.request.query_params.get('entry_id')
        if entry_id:
            queryset = queryset.filter(entry__id=entry_id)
        
        return queryset.order_by('-entry__date', '-entry__id', 'id')

class ARInvoiceViewSet(viewsets.ModelViewSet):
    serializer_class = ARInvoiceSerializer
    queryset = ARInvoice.objects.select_related('customer', 'currency').prefetch_related('items', 'items__tax_rate')
    
    def update(self, request, *args, **kwargs):
        """Override update to prevent modification after submission for approval or posting"""
        invoice = self.get_object()
        
        # Prevent updates if posted
        if invoice.is_posted:
            return Response(
                {"error": "Posted invoices cannot be modified. Use reversal if needed."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Prevent updates if pending approval or approved
        if invoice.approval_status in ['PENDING_APPROVAL', 'APPROVED']:
            return Response(
                {"error": f"Invoices with status '{invoice.approval_status}' cannot be modified."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return super().update(request, *args, **kwargs)
    
    def partial_update(self, request, *args, **kwargs):
        """Override partial_update with same restrictions as update"""
        return self.update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        """Override destroy to prevent deletion after submission for approval or posting"""
        invoice = self.get_object()
        
        # Prevent deletion if posted
        if invoice.is_posted:
            return Response(
                {"error": "Posted invoices cannot be deleted. Use reversal if needed."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Prevent deletion if pending approval or approved
        if invoice.approval_status in ['PENDING_APPROVAL', 'APPROVED']:
            return Response(
                {"error": f"Invoices with status '{invoice.approval_status}' cannot be deleted."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return super().destroy(request, *args, **kwargs)
    
    @extend_schema(
    responses=inline_serializer(
        name="PostGLResponse",
        fields={
            "already_posted": drf_serializers.BooleanField(),
            "journal": inline_serializer(
                name="JournalEntryInline",
                fields={
                    "id": drf_serializers.IntegerField(),
                    "date": drf_serializers.DateField(),
                    "currency": inline_serializer(
                        name="CurrencyInline",
                        fields={"code": drf_serializers.CharField()}
                    ),
                    "memo": drf_serializers.CharField(allow_null=True, required=False),
                    "posted": drf_serializers.BooleanField(),
                    "lines": drf_serializers.ListField(
                        child=inline_serializer(
                            name="JournalLineInline",
                            fields={
                                "account_code": drf_serializers.CharField(),
                                "account_name": drf_serializers.CharField(),
                                "debit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                                "credit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                            }
                        )
                    ),
                },
            ),
        },
    ),
    examples=[
        OpenApiExample(
            "First time posting — created",
            value={
                "already_posted": False,
                "journal": {
                    "id": 123,
                    "date": "2025-10-01",
                    "currency": {"code": "AED"},
                    "memo": "AR Invoice #INV-001",
                    "posted": True,
                    "lines": [
                        {"account_code": "1100", "account_name": "Accounts Receivable", "debit": 0.00, "credit": 105.00},
                        {"account_code": "4000", "account_name": "Revenue", "debit": 0.00, "credit": 100.00},
                        {"account_code": "2100", "account_name": "VAT Output", "debit": 0.00, "credit": 5.00},
                    ],
                },
            },
            response_only=True,
        ),
        OpenApiExample(
            "Subsequent call — already posted",
            value={
                "already_posted": True,
                "journal": {
                    "id": 123,
                    "date": "2025-10-01",
                    "currency": {"code": "AED"},
                    "memo": "AR Invoice #INV-001",
                    "posted": True,
                    "lines": [
                        {"account_code": "1100", "account_name": "Accounts Receivable", "debit": 0.00, "credit": 105.00},
                        {"account_code": "4000", "account_name": "Revenue", "debit": 0.00, "credit": 100.00},
                        {"account_code": "2100", "account_name": "VAT Output", "debit": 0.00, "credit": 5.00},
                    ],
                },
            },
            response_only=True,
        ),
    ],
)
    @action(detail=True, methods=["post"], url_path="post-gl")
    def post_gl(self, request, pk=None):
        invoice = self.get_object()
        
        # Require approval before posting
        if invoice.approval_status != 'APPROVED':
            return Response(
                {"error": f"Invoice must be APPROVED before posting. Current status: {invoice.approval_status or 'DRAFT'}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            je, created = gl_post_from_ar_balanced(invoice)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            # Log unexpected errors
            import traceback
            traceback.print_exc()
            return Response(
                {"detail": f"Failed to post invoice to GL: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        payload = {"already_posted": not created, "journal": JournalEntryReadSerializer(je).data}
        return Response(payload, status=200 if created or not created else 201)

    @action(detail=True, methods=["post"], url_path="submit-for-approval")
    def submit_for_approval(self, request, pk=None):
        """Submit AR invoice for approval"""
        invoice = self.get_object()
        
        # Check if invoice is in draft status
        if invoice.is_posted:
            return Response(
                {"error": "Posted invoices cannot be submitted for approval"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if invoice.is_cancelled:
            return Response(
                {"error": "Cancelled invoices cannot be submitted for approval"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if already submitted
        if invoice.approval_status == 'PENDING_APPROVAL':
            return Response(
                {"error": "Invoice is already pending approval"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if invoice.approval_status == 'APPROVED':
            return Response(
                {"error": "Invoice is already approved"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Calculate invoice total
        totals = ar_totals(invoice)
        invoice_total = totals['total']
        
        # Create approval record
        approval = InvoiceApproval.objects.create(
            invoice_id=invoice.id,
            invoice_type='AR',
            submitted_by=request.data.get('submitted_by', 'System'),
            submitted_at=timezone.now(),
            status='PENDING_APPROVAL'
        )
        
        # Update invoice approval status
        invoice.approval_status = 'PENDING_APPROVAL'
        invoice.save()
        
        return Response({
            "message": "Invoice submitted for approval successfully",
            "approval_id": approval.id,
            "invoice_id": invoice.id,
            "invoice_number": invoice.number,
            "invoice_total": str(invoice_total),
            "currency": invoice.currency.code,
            "approval_status": invoice.approval_status
        }, status=status.HTTP_201_CREATED)



class APInvoiceViewSet(viewsets.ModelViewSet):
    serializer_class = APInvoiceSerializer
    queryset = APInvoice.objects.select_related('supplier', 'currency').prefetch_related('items', 'items__tax_rate')
    
    def update(self, request, *args, **kwargs):
        """Override update to prevent modification after submission for approval or posting"""
        invoice = self.get_object()
        
        # Prevent updates if posted
        if invoice.is_posted:
            return Response(
                {"error": "Posted invoices cannot be modified. Use reversal if needed."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Prevent updates if pending approval or approved
        if invoice.approval_status in ['PENDING_APPROVAL', 'APPROVED']:
            return Response(
                {"error": f"Invoices with status '{invoice.approval_status}' cannot be modified."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return super().update(request, *args, **kwargs)
    
    def partial_update(self, request, *args, **kwargs):
        """Override partial_update with same restrictions as update"""
        return self.update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        """Override destroy to prevent deletion after submission for approval or posting"""
        invoice = self.get_object()
        
        # Prevent deletion if posted
        if invoice.is_posted:
            return Response(
                {"error": "Posted invoices cannot be deleted. Use reversal if needed."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Prevent deletion if pending approval or approved
        if invoice.approval_status in ['PENDING_APPROVAL', 'APPROVED']:
            return Response(
                {"error": f"Invoices with status '{invoice.approval_status}' cannot be deleted."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return super().destroy(request, *args, **kwargs)
    
    @extend_schema(
    responses=inline_serializer(
        name="APPostGLResponse",
        fields={
            "already_posted": drf_serializers.BooleanField(),
            "journal": inline_serializer(
                name="APJournalEntryInline",
                fields={
                    "id": drf_serializers.IntegerField(),
                    "date": drf_serializers.DateField(),
                    "currency": inline_serializer(
                        name="APCurrencyInline",
                        fields={"code": drf_serializers.CharField()}
                    ),
                    "memo": drf_serializers.CharField(allow_null=True, required=False),
                    "posted": drf_serializers.BooleanField(),
                    "lines": drf_serializers.ListField(
                        child=inline_serializer(
                            name="APJournalLineInline",
                            fields={
                                "account_code": drf_serializers.CharField(),
                                "account_name": drf_serializers.CharField(),
                                "debit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                                "credit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                            }
                        )
                    ),
                },
            ),
        },
    ),
    examples=[
        OpenApiExample(
            "First time posting — created",
            value={
                "already_posted": False,
                "journal": {
                    "id": 456,
                    "date": "2025-10-01",
                    "currency": {"code": "AED"},
                    "memo": "AP Invoice #INV-001",
                    "posted": True,
                    "lines": [
                        {"account_code": "2000", "account_name": "Accounts Payable", "debit": 115.00, "credit": 0.00},
                        {"account_code": "5000", "account_name": "Expenses", "debit": 0.00, "credit": 100.00},
                        {"account_code": "2150", "account_name": "VAT Input", "debit": 0.00, "credit": 15.00},
                    ],
                },
            },
            response_only=True,
        ),
        OpenApiExample(
            "Subsequent call — already posted",
            value={
                "already_posted": True,
                "journal": {
                    "id": 456,
                    "date": "2025-10-01",
                    "currency": {"code": "AED"},
                    "memo": "AP Invoice #INV-001",
                    "posted": True,
                    "lines": [
                        {"account_code": "2000", "account_name": "Accounts Payable", "debit": 115.00, "credit": 0.00},
                        {"account_code": "5000", "account_name": "Expenses", "debit": 0.00, "credit": 100.00},
                        {"account_code": "2150", "account_name": "VAT Input", "debit": 0.00, "credit": 15.00},
                    ],
                },
            },
            response_only=True,
        ),
    ],
)
    @action(detail=True, methods=["post"], url_path="post-gl")
    def post_gl(self, request, pk=None):
        invoice = self.get_object()
        
        # Require approval before posting
        if invoice.approval_status != 'APPROVED':
            return Response(
                {"error": f"Invoice must be APPROVED before posting. Current status: {invoice.approval_status or 'DRAFT'}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            je, created = gl_post_from_ap_balanced(invoice)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            # Log unexpected errors
            import traceback
            traceback.print_exc()
            return Response(
                {"detail": f"Failed to post invoice to GL: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        payload = {"already_posted": not created, "journal": JournalEntryReadSerializer(je).data}
        return Response(payload, status=200 if created or not created else 201)

    @action(detail=True, methods=["post"], url_path="submit-for-approval")
    def submit_for_approval(self, request, pk=None):
        """Submit AP invoice for approval"""
        invoice = self.get_object()
        
        # Check if invoice is in draft status
        if invoice.is_posted:
            return Response(
                {"error": "Posted invoices cannot be submitted for approval"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if invoice.is_cancelled:
            return Response(
                {"error": "Cancelled invoices cannot be submitted for approval"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if already submitted
        if invoice.approval_status == 'PENDING_APPROVAL':
            return Response(
                {"error": "Invoice is already pending approval"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if invoice.approval_status == 'APPROVED':
            return Response(
                {"error": "Invoice is already approved"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Calculate invoice total
        totals = ap_totals(invoice)
        invoice_total = totals['total']
        
        # Create approval record
        approval = InvoiceApproval.objects.create(
            invoice_id=invoice.id,
            invoice_type='AP',
            submitted_by=request.data.get('submitted_by', 'System'),
            submitted_at=timezone.now(),
            status='PENDING_APPROVAL'
        )
        
        # Update invoice approval status
        invoice.approval_status = 'PENDING_APPROVAL'
        invoice.save()
        
        return Response({
            "message": "Invoice submitted for approval successfully",
            "approval_id": approval.id,
            "invoice_id": invoice.id,
            "invoice_number": invoice.number,
            "invoice_total": str(invoice_total),
            "currency": invoice.currency.code,
            "approval_status": invoice.approval_status
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'], url_path='three-way-match')
    def three_way_match(self, request, pk=None):
        """
        Perform 3-way match on AP invoice linked to GRN.
        Compares: PO → GRN → Invoice (quantities, prices, amounts)
        """
        invoice = self.get_object()
        
        # Check if invoice has GRN link
        if not invoice.goods_receipt:
            return Response(
                {'error': '3-way match requires invoice to be linked to a GRN'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not invoice.po_header:
            return Response(
                {'error': '3-way match requires invoice to be linked to a PO'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Perform 3-way match
        try:
            from ap.services import perform_three_way_match
            result = perform_three_way_match(invoice)
            
            return Response({
                'message': 'Three-way match completed',
                'invoice_id': invoice.id,
                'invoice_number': invoice.number,
                'match_status': result['status'],
                'variances': result['variances'],
                'total_variance_amount': str(result['total_variance_amount']),
                'notes': result['notes'],
                'performed_at': invoice.match_performed_at.isoformat() if invoice.match_performed_at else None
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Failed to perform 3-way match: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='grns-needing-invoice')
    def grns_needing_invoice(self, request):
        """
        Get list of GRNs (Goods Receipt Notes) that need AP invoicing.
        Returns COMPLETED GRNs that don't have associated AP invoices yet.
        """
        from procurement.receiving.models import GoodsReceipt
        from procurement.receiving.serializers import GoodsReceiptListSerializer
        
        # Get all COMPLETED GRNs
        grns = GoodsReceipt.objects.filter(
            status='COMPLETED'
        ).exclude(
            # Exclude GRNs that already have non-cancelled AP invoices
            id__in=APInvoice.objects.filter(
                goods_receipt__isnull=False,
                is_cancelled=False
            ).values_list('goods_receipt_id', flat=True).distinct()
        ).select_related(
            'supplier', 'po_header', 'warehouse'
        ).order_by('-receipt_date')
        
        # Apply filters
        supplier_id = request.query_params.get('supplier')
        if supplier_id:
            grns = grns.filter(supplier_id=supplier_id)
        
        po_number = request.query_params.get('po_number')
        if po_number:
            grns = grns.filter(po_header__po_number__icontains=po_number)
        
        serializer = GoodsReceiptListSerializer(grns, many=True)
        return Response({
            'count': grns.count(),
            'results': serializer.data
        })


class ARPaymentViewSet(viewsets.ModelViewSet):
    serializer_class = ARPaymentSerializer
    queryset = ARPayment.objects.all()

    def perform_create(self, serializer):
        payment = serializer.save()  # creates ARPayment
        je, created, closed = post_ar_payment(payment)
        # We return enriched response
        self._last_payload = {
            "journal": JournalEntryReadSerializer(je).data,
            "already_posted": not created,
            "invoice_closed": closed,
        }

    def create(self, request, *args, **kwargs):
        resp = super().create(request, *args, **kwargs)
        # merge with normal serializer output if you like; simplest: replace with posting payload
        return Response(self._last_payload, status=201)

    def perform_update(self, serializer):
        """Reverse old GL entry and create new one when payment is updated"""
        payment = self.get_object()
        old_gl_journal = payment.gl_journal
        old_amount = payment.amount
        
        # Save the updated payment
        payment = serializer.save()
        
        # If GL journal exists and amount changed, reverse and repost
        if old_gl_journal and old_amount != payment.amount:
            reverse_journal(old_gl_journal)
            je, created, closed = post_ar_payment(payment)
            self._last_payload = {
                "journal": JournalEntryReadSerializer(je).data,
                "reversed_journal": old_gl_journal.id,
                "invoice_closed": closed,
            }
        else:
            self._last_payload = {"status": "updated", "no_gl_change": True}

    def update(self, request, *args, **kwargs):
        resp = super().update(request, *args, **kwargs)
        return Response(self._last_payload, status=200)

    def perform_destroy(self, instance):
        """Reverse GL entry when payment is deleted"""
        if instance.gl_journal:
            reverse_journal(instance.gl_journal)
        instance.delete()

    @action(detail=True, methods=["post"])
    def reconcile(self, request, pk=None):
        payment = self.get_object()
        payment.reconciled = True
        payment.reconciliation_ref = request.data.get("reconciliation_ref", "")
        payment.reconciled_at = timezone.now().date()
        payment.save()
        return Response({"status": "reconciled"}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    def unreconcile(self, request, pk=None):
        payment = self.get_object()
        payment.reconciled = False
        payment.reconciliation_ref = ""
        payment.reconciled_at = None
        payment.save()
        return Response({"status": "unreconciled"}, status=status.HTTP_200_OK)
@extend_schema_view(
    create=extend_schema(
        examples=[
            OpenApiExample(
                "Simple AP payment",
                value={"invoice": 456, "date": "2025-10-02", "amount": 115.00, "bank_account": 1},
                request_only=True,
            )
        ],
        responses=inline_serializer(
            name="APPaymentResponse",
            fields={
                "already_posted": drf_serializers.BooleanField(),
                "invoice_closed": drf_serializers.BooleanField(),
                "journal": inline_serializer(
                    name="APPaymentJournalEntryInline",
                    fields={
                        "id": drf_serializers.IntegerField(),
                        "date": drf_serializers.DateField(),
                        "currency": inline_serializer(
                            name="APPaymentCurrencyInline",
                            fields={"code": drf_serializers.CharField()}
                        ),
                        "memo": drf_serializers.CharField(allow_null=True, required=False),
                        "posted": drf_serializers.BooleanField(),
                        "lines": drf_serializers.ListField(
                            child=inline_serializer(
                                name="APPaymentJournalLineInline",
                                fields={
                                    "account_code": drf_serializers.CharField(),
                                    "account_name": drf_serializers.CharField(),
                                    "debit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                                    "credit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                                }
                            )
                        ),
                    },
                ),
            },
        ),
    )
)
class APPaymentViewSet(viewsets.ModelViewSet):
    serializer_class = APPaymentSerializer
    queryset = APPayment.objects.all()

    def perform_create(self, serializer):
        payment = serializer.save()
        je, created, closed = post_ap_payment(payment)
        self._last_payload = {
            "journal": JournalEntryReadSerializer(je).data,
            "already_posted": not created,
            "invoice_closed": closed,
        }
    @action(detail=True, methods=["post"])
    def reconcile(self, request, pk=None):
        payment = self.get_object()
        payment.reconciled = True
        payment.reconciliation_ref = request.data.get("reconciliation_ref", "")
        payment.reconciled_at = timezone.now().date()
        payment.save()
        return Response({"status": "reconciled"}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    def unreconcile(self, request, pk=None):
        payment = self.get_object()
        payment.reconciled = False
        payment.reconciliation_ref = ""
        payment.reconciled_at = None
        payment.save()
        return Response({"status": "unreconciled"}, status=status.HTTP_200_OK)
    def create(self, request, *args, **kwargs):
        resp = super().create(request, *args, **kwargs)
        return Response(self._last_payload, status=201)

    def perform_update(self, serializer):
        """Reverse old GL entry and create new one when payment is updated"""
        payment = self.get_object()
        old_gl_journal = payment.gl_journal
        old_amount = payment.amount
        
        # Save the updated payment
        payment = serializer.save()
        
        # If GL journal exists and amount changed, reverse and repost
        if old_gl_journal and old_amount != payment.amount:
            reverse_journal(old_gl_journal)
            je, created, closed = post_ap_payment(payment)
            self._last_payload = {
                "journal": JournalEntryReadSerializer(je).data,
                "reversed_journal": old_gl_journal.id,
                "invoice_closed": closed,
            }
        else:
            self._last_payload = {"status": "updated", "no_gl_change": True}

    def update(self, request, *args, **kwargs):
        resp = super().update(request, *args, **kwargs)
        return Response(self._last_payload, status=200)

    def perform_destroy(self, instance):
        """Reverse GL entry when payment is deleted"""
        if instance.gl_journal:
            reverse_journal(instance.gl_journal)
        instance.delete()

# ---- Seed VAT presets ----
class SeedVATPresets(APIView):
    # permission_classes = [IsAdminUser]  # Disabled for testing
    def post(self, request):
        s = SeedVATRequestSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        ids = seed_vat_presets(effective_from=s.validated_data.get("effective_from"))
        return Response({"created_ids": ids}, status=201 if ids else 200)

# ---- List tax rates by country ----
class ListTaxRates(APIView):
    def get(self, request):
        country = request.GET.get("country")
        qs = TaxRate.objects.all()
        if country:
            qs = qs.filter(country=country)
        data = [{
            "id": t.id, 
            "name": t.name, 
            "rate": float(t.rate), 
            "country": t.country,
            "category": t.category, 
            "code": t.code, 
            "effective_from": t.effective_from.isoformat() if t.effective_from else None,
            "is_active": t.is_active  # Added is_active field
        } for t in qs.order_by("country","category","rate")]
        return Response(data)

# ---- Corporate tax accrual ----
class CorporateTaxAccrual(APIView):
   # permission_classes = [IsAdminUser]  # Disabled for testing
   def post(self, request):
       s = CorpTaxAccrualRequestSerializer(data=request.data)
       s.is_valid(raise_exception=True)
       country = s.validated_data["country"]
       df = s.validated_data["date_from"]; dt = s.validated_data["date_to"]
       org_id = request.data.get("org_id")
       allow_override = bool(request.data.get("override", False))
       filing, je, meta = accrue_corporate_tax_with_filing(country, df, dt, org_id, allow_override=allow_override)
       if je is None:
           return Response({"created": False, "meta": meta}, status=200)
       return Response({
           "created": True, "filing_id": filing.id,
           "journal": JournalEntryReadSerializer(je).data, "meta": meta
       }, status=201)
   
class CorporateTaxFile(APIView):
   # permission_classes = [IsAdminUser]  # Disabled for testing
   
   def get(self, request, filing_id:int):
       """Get filing details - use GET /api/tax/corporate-filing/{id}/ instead"""
       try:
           filing = CorporateTaxFiling.objects.get(pk=filing_id)
           return Response({
               "message": f"Use GET /api/tax/corporate-filing/{filing_id}/ to retrieve filing details",
               "current_status": filing.status,
               "can_file": filing.status == "ACCRUED",
               "filed_at": filing.filed_at.isoformat() if filing.filed_at else None,
           }, status=200)
       except CorporateTaxFiling.DoesNotExist:
           return Response({"detail": f"CorporateTaxFiling with id {filing_id} does not exist."}, status=status.HTTP_404_NOT_FOUND)
   
   def post(self, request, filing_id:int):
       try:
           f = file_corporate_tax(filing_id)
           return Response({"filing_id": f.id, "status": f.status, "filed_at": f.filed_at}, status=200)
       except CorporateTaxFiling.DoesNotExist:
           return Response({"detail": f"CorporateTaxFiling with id {filing_id} does not exist."}, status=status.HTTP_404_NOT_FOUND)
       except ValueError as e:
           # Try to get current status for better error message
           try:
               filing = CorporateTaxFiling.objects.get(pk=filing_id)
               return Response({
                   "detail": str(e),
                   "current_status": filing.status,
                   "required_status": "ACCRUED"
               }, status=status.HTTP_400_BAD_REQUEST)
           except CorporateTaxFiling.DoesNotExist:
               return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class CorporateTaxReverse(APIView):
   # permission_classes = [IsAdminUser]  # Disabled for testing
   
   def get(self, request, filing_id:int):
       """Get filing details - use GET /api/tax/corporate-filing/{id}/ instead"""
       try:
           filing = CorporateTaxFiling.objects.get(pk=filing_id)
           return Response({
               "message": f"Use GET /api/tax/corporate-filing/{filing_id}/ to retrieve filing details",
               "current_status": filing.status,
               "can_reverse": filing.status in ["ACCRUED", "FILED"] and filing.accrual_journal_id is not None,
               "reason": self._get_cannot_reverse_reason(filing),
           }, status=200)
       except CorporateTaxFiling.DoesNotExist:
           return Response({"detail": f"CorporateTaxFiling with id {filing_id} does not exist."}, status=status.HTTP_404_NOT_FOUND)
   
   def _get_cannot_reverse_reason(self, filing):
       """Helper to explain why reversal cannot be performed"""
       if filing.status == "REVERSED":
           return "Filing is already reversed"
       if filing.status == "FILED":
           return "Filing is locked (FILED status). Cannot reverse without unfile/override."
       if not filing.accrual_journal_id:
           return "No accrual journal to reverse"
       return None
   
   def post(self, request, filing_id:int):
       try:
           rev = reverse_corporate_tax_filing(filing_id)
           return Response({"reversal_journal": JournalEntryReadSerializer(rev).data}, status=201)
       except CorporateTaxFiling.DoesNotExist:
           return Response({"detail": f"CorporateTaxFiling with id {filing_id} does not exist."}, status=status.HTTP_404_NOT_FOUND)
       except ValueError as e:
           # Try to get current status for better error message
           try:
               filing = CorporateTaxFiling.objects.get(pk=filing_id)
               return Response({
                   "detail": str(e),
                   "current_status": filing.status,
                   "has_accrual_journal": filing.accrual_journal_id is not None,
                   "already_reversed": filing.reversal_journal_id is not None
               }, status=status.HTTP_400_BAD_REQUEST)
           except CorporateTaxFiling.DoesNotExist:
               return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class CorporateTaxFilingDetail(APIView):
   # permission_classes = [IsAdminUser]  # Disabled for testing
   def get(self, request, filing_id:int):
       try:
           filing = CorporateTaxFiling.objects.get(pk=filing_id)
           data = {
               "id": filing.id,
               "country": filing.country,
               "period_start": filing.period_start.isoformat(),
               "period_end": filing.period_end.isoformat(),
               "organization_id": filing.organization_id,
               "status": filing.status,
               "filed_at": filing.filed_at.isoformat() if filing.filed_at else None,
               "notes": filing.notes,
               "accrual_journal_id": filing.accrual_journal_id,
               "reversal_journal_id": filing.reversal_journal_id,
               "accrual_journal": JournalEntryReadSerializer(filing.accrual_journal).data if filing.accrual_journal else None,
               "reversal_journal": JournalEntryReadSerializer(filing.reversal_journal).data if filing.reversal_journal else None,
           }
           return Response(data, status=200)
       except CorporateTaxFiling.DoesNotExist:
           return Response({"detail": f"CorporateTaxFiling with id {filing_id} does not exist."}, status=status.HTTP_404_NOT_FOUND)
   

class CorporateTaxBreakdown(APIView):
   def get(self, request):
       country = request.GET.get("country")
       df = request.GET.get("date_from"); dt = request.GET.get("date_to")
       org_id = request.GET.get("org_id")
       fmt = (request.GET.get("format") or "json").lower()
       file_type = request.GET.get("file_type", "").lower()
       # recompute using same logic as accrual (but don’t post)
       from datetime import datetime
       df_d = datetime.fromisoformat(df).date() if df else None
       dt_d = datetime.fromisoformat(dt).date() if dt else None
       # collect rows
       qs = JournalLine.objects.select_related("entry","account") \
            .filter(entry__posted=True)
       if df_d: qs = qs.filter(entry__date__gte=df_d)
       if dt_d: qs = qs.filter(entry__date__lte=dt_d)
       qs = _with_org_filter(qs, org_id)
       rows = []
       income = Decimal("0"); expense = Decimal("0")
       for ln in qs:
           kind = "OTHER"
           delta = Decimal("0")
           # Handle both short codes (IN, EX) and full names (INCOME, EXPENSE)
           acc_type = ln.account.type.upper()
           if acc_type in ("INCOME", "IN"):
               kind = "INCOME"
               delta = (Decimal(ln.credit) - Decimal(ln.debit))
               income += delta
           elif acc_type in ("EXPENSE", "EX"):
               kind = "EXPENSE"
               delta = (Decimal(ln.debit) - Decimal(ln.credit))
               expense += delta
           rows.append({
               "date": ln.entry.date.isoformat() if ln.entry.date else None,
               "journal_id": ln.entry.id,
               "account_code": ln.account.code,
               "account_name": ln.account.name,
               "type": kind,
               "delta": float(q2(delta)),
               "debit": float(q2(ln.debit)),
               "credit": float(q2(ln.credit)),
           })
       profit = q2(income - expense)
       meta = {
           "country": country,
           "date_from": df, "date_to": dt, "org_id": org_id,
           "income": float(q2(income)), "expense": float(q2(expense)),
           "profit": float(profit)
       }
       # CSV Export
       if fmt == "csv" or file_type == "csv":
           from django.http import HttpResponse; import csv
           resp = HttpResponse(content_type="text/csv; charset=utf-8")
           resp["Content-Disposition"] = 'attachment; filename="corp_tax_breakdown.csv"'
           resp.write('\ufeff')  # Excel-friendly BOM
           w = csv.writer(resp)
           w.writerow(["Date","Journal","Account Code","Account Name","Type","Delta","Debit","Credit"])
           for r in rows: w.writerow([r["date"], r["journal_id"], r["account_code"], r["account_name"], r["type"], f'{r["delta"]:.2f}', f'{r["debit"]:.2f}', f'{r["credit"]:.2f}'])
           # footer
           w.writerow([]); w.writerow(["Income", f'{meta["income"]:.2f}'])
           w.writerow(["Expense", f'{meta["expense"]:.2f}'])
           w.writerow(["Profit", f'{meta["profit"]:.2f}'])
           return resp
       # XLSX Export
       if fmt in ("xlsx","excel") or file_type == "xlsx":
           if not OPENPYXL_OK:
               return HttpResponse("openpyxl not installed", status=400)
           from io import BytesIO
           wb = Workbook(); ws = wb.active; ws.title="CorpTax Breakdown"
           ws.append(["Date","Journal","Account Code","Account Name","Type","Delta","Debit","Credit"])
           for r in rows: ws.append([r["date"], r["journal_id"], r["account_code"], r["account_name"], r["type"], r["delta"], r["debit"], r["credit"]])
           ws.append([]); ws.append(["Income", meta["income"]]); ws.append(["Expense", meta["expense"]]); ws.append(["Profit", meta["profit"]])
           bio = BytesIO(); wb.save(bio); bio.seek(0)
           from django.http import HttpResponse
           resp = HttpResponse(bio.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
           resp["Content-Disposition"] = 'attachment; filename="corp_tax_breakdown.xlsx"'
           return resp
       
       # Default JSON response with download links (like Trail Balance)
       base_url = request.build_absolute_uri(request.path)
       query_params = []
       if country:
           query_params.append(f"country={country}")
       if df:
           query_params.append(f"date_from={df}")
       if dt:
           query_params.append(f"date_to={dt}")
       if org_id:
           query_params.append(f"org_id={org_id}")
       
       query_string = "&".join(query_params)
       
       response_data = {
           "meta": meta,
           "rows": rows,
           "download_links": {
               "xlsx": f"{base_url}?file_type=xlsx" + (f"&{query_string}" if query_string else ""),
               "csv": f"{base_url}?file_type=csv" + (f"&{query_string}" if query_string else "")
           }
       }
       
       return Response(response_data, status=200)


# ========== DEPRECATED LEGACY INVOICE API REMOVED ==========
# InvoiceViewSet was removed as part of consolidation to AR/AP specific invoice models.
# Use ARInvoiceViewSet or APInvoiceViewSet instead.


# ========== FX (Foreign Exchange) API Endpoints ==========

from core.models import ExchangeRate, FXGainLossAccount, Currency
from .serializers import (
    ExchangeRateSerializer, ExchangeRateCreateSerializer,
    FXGainLossAccountSerializer, CurrencyConversionRequestSerializer
)
from .fx_services import (
    get_exchange_rate, convert_amount, get_base_currency,
    create_exchange_rate
)


class ExchangeRateViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing exchange rates.
    Supports CRUD operations on exchange rates.
    """
    queryset = ExchangeRate.objects.select_related('from_currency', 'to_currency').all()
    serializer_class = ExchangeRateSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by currency codes or IDs
        from_currency = self.request.query_params.get('from_currency')
        to_currency = self.request.query_params.get('to_currency')
        rate_type = self.request.query_params.get('rate_type')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        
        if from_currency:
            # Support both currency ID and code
            if from_currency.isdigit():
                queryset = queryset.filter(from_currency__id=int(from_currency))
            else:
                queryset = queryset.filter(from_currency__code=from_currency)
        if to_currency:
            # Support both currency ID and code
            if to_currency.isdigit():
                queryset = queryset.filter(to_currency__id=int(to_currency))
            else:
                queryset = queryset.filter(to_currency__code=to_currency)
        if rate_type:
            queryset = queryset.filter(rate_type=rate_type)
        if date_from and date_to:
            # For a specific date range, find the most recent rate on or before date_to
            queryset = queryset.filter(rate_date__lte=date_to).order_by('-rate_date')
        elif date_from:
            queryset = queryset.filter(rate_date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(rate_date__lte=date_to)
        
        return queryset


class CurrencyConvertView(APIView):
    """
    Convert an amount from one currency to another using exchange rates.
    POST with: amount, from_currency_code, to_currency_code, rate_date, rate_type (optional)
    """
    def post(self, request):
        serializer = CurrencyConversionRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        amount = data['amount']
        from_code = data['from_currency_code']
        to_code = data['to_currency_code']
        rate_date = data['rate_date']
        rate_type = data.get('rate_type', 'SPOT')
        
        try:
            from_currency = Currency.objects.get(code=from_code)
            to_currency = Currency.objects.get(code=to_code)
            
            rate = get_exchange_rate(from_currency, to_currency, rate_date, rate_type)
            converted_amount = convert_amount(amount, from_currency, to_currency, rate_date, rate_type)
            
            return Response({
                'from_currency': from_code,
                'to_currency': to_code,
                'original_amount': float(amount),
                'exchange_rate': float(rate),
                'converted_amount': float(converted_amount),
                'rate_date': rate_date.isoformat(),
                'rate_type': rate_type
            }, status=200)
        except Currency.DoesNotExist as e:
            return Response({'detail': str(e)}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CreateExchangeRateView(APIView):
    """
    Create or update an exchange rate using currency codes.
    POST with: from_currency_code, to_currency_code, rate, rate_date, rate_type (optional), source (optional)
    """
    def post(self, request):
        serializer = ExchangeRateCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        
        try:
            rate_obj = create_exchange_rate(
                from_currency_code=data['from_currency_code'],
                to_currency_code=data['to_currency_code'],
                rate=data['rate'],
                rate_date=data['rate_date'],
                rate_type=data.get('rate_type', 'SPOT'),
                source=data.get('source', '')
            )
            
            return Response(ExchangeRateSerializer(rate_obj).data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class FXGainLossAccountViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing FX Gain/Loss account configurations.
    """
    queryset = FXGainLossAccount.objects.select_related('account').all()
    serializer_class = FXGainLossAccountSerializer


class BaseCurrencyView(APIView):
    """
    Get the base/home currency for the company.
    """
    def get(self, request):
        try:
            base_currency = get_base_currency()
            return Response({
                'code': base_currency.code,
                'name': base_currency.name,
                'symbol': base_currency.symbol
            }, status=200)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)