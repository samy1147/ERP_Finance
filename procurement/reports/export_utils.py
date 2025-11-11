"""
Export Utilities for Procurement Reports

Utilities for exporting data to CSV and XLSX formats.
"""

import csv
import io
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class CSVExporter:
    """Export data to CSV format."""
    
    @staticmethod
    def export_to_response(data, filename, headers=None):
        """
        Export data to CSV HttpResponse.
        
        Args:
            data: List of dictionaries or QuerySet
            filename: Output filename (without .csv extension)
            headers: Optional list of header names
        
        Returns:
            HttpResponse with CSV content
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        if not data:
            return response
        
        # Convert QuerySet to list of dicts if needed
        if hasattr(data, 'values'):
            data = list(data.values())
        
        # Get headers from first row if not provided
        if headers is None:
            headers = list(data[0].keys()) if data else []
        
        writer = csv.DictWriter(response, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        
        return response
    
    @staticmethod
    def export_to_string(data, headers=None):
        """
        Export data to CSV string.
        
        Args:
            data: List of dictionaries
            headers: Optional list of header names
        
        Returns:
            CSV string
        """
        if not data:
            return ""
        
        output = io.StringIO()
        
        if headers is None:
            headers = list(data[0].keys()) if data else []
        
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        
        return output.getvalue()


class ExcelExporter:
    """Export data to Excel (XLSX) format."""
    
    @staticmethod
    def export_to_response(data, filename, headers=None, sheet_name='Sheet1', 
                          title=None, apply_formatting=True):
        """
        Export data to Excel HttpResponse with formatting.
        
        Args:
            data: List of dictionaries or QuerySet
            filename: Output filename (without .xlsx extension)
            headers: Optional list of header names
            sheet_name: Name of the Excel sheet
            title: Optional title for the sheet
            apply_formatting: Whether to apply formatting
        
        Returns:
            HttpResponse with Excel content
        """
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Convert QuerySet to list of dicts if needed
        if hasattr(data, 'values'):
            data = list(data.values())
        
        if not data:
            wb = Workbook()
            wb.save(response)
            return response
        
        # Get headers from first row if not provided
        if headers is None:
            headers = list(data[0].keys()) if data else []
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        current_row = 1
        
        # Add title if provided
        if title and apply_formatting:
            ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row = 2
        
        # Write headers
        if apply_formatting:
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header.replace('_', ' ').title()
            if apply_formatting:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 1
        
        # Write data
        for row_data in data:
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                value = row_data.get(header, '')
                
                # Format datetime objects
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                
                cell.value = value
                
                # Apply number formatting for numeric values
                if apply_formatting and isinstance(value, (int, float)):
                    if isinstance(value, float) and value != int(value):
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '#,##0'
            
            current_row += 1
        
        # Auto-size columns
        if apply_formatting:
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(str(header))
                
                for row in ws[column_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response
    
    @staticmethod
    def export_multi_sheet(sheets_data, filename, apply_formatting=True):
        """
        Export multiple sheets to single Excel file.
        
        Args:
            sheets_data: List of tuples (sheet_name, data, headers, title)
            filename: Output filename
            apply_formatting: Whether to apply formatting
        
        Returns:
            HttpResponse with Excel content
        """
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for sheet_name, data, headers, title in sheets_data:
            ws = wb.create_sheet(title=sheet_name)
            
            # Convert QuerySet to list of dicts if needed
            if hasattr(data, 'values'):
                data = list(data.values())
            
            if not data:
                continue
            
            # Get headers from first row if not provided
            if headers is None:
                headers = list(data[0].keys()) if data else []
            
            current_row = 1
            
            # Add title if provided
            if title and apply_formatting:
                ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
                title_cell = ws['A1']
                title_cell.value = title
                title_cell.font = Font(size=14, bold=True)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row = 2
            
            # Write headers
            if apply_formatting:
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header.replace('_', ' ').title()
                if apply_formatting:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
            
            # Write data
            for row_data in data:
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    value = row_data.get(header, '')
                    
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    
                    cell.value = value
                    
                    if apply_formatting and isinstance(value, (int, float)):
                        if isinstance(value, float) and value != int(value):
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
                
                current_row += 1
            
            # Auto-size columns
            if apply_formatting:
                for col_num, header in enumerate(headers, 1):
                    column_letter = get_column_letter(col_num)
                    max_length = len(str(header))
                    
                    for row in ws[column_letter]:
                        try:
                            if len(str(row.value)) > max_length:
                                max_length = len(str(row.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response


def prepare_export_data(queryset, fields):
    """
    Prepare queryset data for export with proper field extraction.
    
    Args:
        queryset: Django QuerySet
        fields: List of field names or tuples (field_name, display_name)
    
    Returns:
        List of dictionaries ready for export
    """
    data = []
    
    for obj in queryset:
        row = {}
        for field in fields:
            if isinstance(field, tuple):
                field_name, display_name = field
            else:
                field_name = field
                display_name = field
            
            # Handle nested fields (e.g., 'supplier.name')
            if '.' in field_name:
                parts = field_name.split('.')
                value = obj
                for part in parts:
                    value = getattr(value, part, None)
                    if value is None:
                        break
            else:
                value = getattr(obj, field_name, None)
            
            # Handle special cases
            if callable(value):
                value = value()
            elif hasattr(value, 'all'):  # ManyToMany field
                value = ', '.join(str(v) for v in value.all())
            
            row[display_name] = value
        
        data.append(row)
    
    return data
