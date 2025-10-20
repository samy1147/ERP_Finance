
from rest_framework import viewsets, status
from rest_framework.permissions import IsAdminUser
from rest_framework.decorators import api_view, permission_classes,action
from rest_framework.response import Response
from rest_framework.views import APIView
from datetime import datetime, date
from django.utils import timezone
from core.models import TaxRate, Currency
from django.http import HttpResponse
import csv
from django.db.models import Sum, F, Q
from io import BytesIO
from decimal import Decimal
from .models import Invoice , Account, JournalEntry, JournalLine, BankAccount
from ar.models import ARInvoice, ARPayment
from ap.models import APInvoice, APPayment
from .serializers import (InvoiceSerializer,AccountSerializer,JournalLineSerializer, JournalEntrySerializer,ARInvoiceSerializer, ARPaymentSerializer,APInvoiceSerializer, APPaymentSerializer,JournalEntryReadSerializer, BankAccountSerializer,SeedVATRequestSerializer, CorpTaxAccrualRequestSerializer)
from .services import (
    reverse_posted_invoice,post_invoice,post_entry,
    gl_post_from_ar_balanced, gl_post_from_ap_balanced,
    post_ar_payment, post_ap_payment,
    reverse_journal, seed_vat_presets, accrue_corporate_tax,
    q2, _with_org_filter
)
from .models import CorporateTaxFiling
from .services import (
    resolve_tax_rate_for_date,
    accrue_corporate_tax_with_filing, reverse_corporate_tax_filing, file_corporate_tax,
    # reuse accrue_corporate_tax if you need raw calc without filing
)
try:
    from openpyxl import Workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


from .services import build_trial_balance, build_ar_aging, build_ap_aging


from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiExample, inline_serializer
from rest_framework import serializers as drf_serializers

@extend_schema(
    parameters=[],
    description="Trial Balance report. Supports ?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD and format choices.",
    responses={200: None}
)
class TrialBalanceReport(APIView):
    def get(self, request):
        df = request.GET.get("date_from")
        dt = request.GET.get("date_to")
        fmt = (request.GET.get("format") or "json").lower()
        file_type = request.GET.get("file_type", "").lower()

        # Force list to avoid generator exhaustion
        rows = list(build_trial_balance(df, dt))

        if fmt == "csv" or file_type == "csv":
            # CSV uses built-in csv module, no external library needed
            resp = HttpResponse(content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="trial_balance.csv"'
            resp.write('\ufeff')  # Excel-friendly BOM
            w = csv.writer(resp)
            w.writerow(["Account Code", "Account Name", "Debit", "Credit"])
            for r in rows:
                w.writerow([
                    r.get("code", ""),
                    r.get("name", ""),
                    f'{r.get("debit", 0):.2f}',
                    f'{r.get("credit", 0):.2f}'
                ])
            return resp

        if fmt in ("xlsx", "excel", "xlsm") or file_type == "xlsx":
            if not OPENPYXL_OK:
               return HttpResponse("openpyxl not installed", status=400)
            wb = Workbook()
            ws = wb.active
            ws.title = "Trial Balance"
            ws.append(["Account Code", "Account Name", "Debit", "Credit"])
            for r in rows:
                ws.append([
                    r.get("code", ""),
                    r.get("name", ""),
                    float(r.get("debit", 0)),
                    float(r.get("credit", 0))
                ])

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            resp = HttpResponse(
                bio.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = 'attachment; filename="trial_balance.xlsx"'
            return resp

        # Default JSON response with download links
        base_url = request.build_absolute_uri(request.path)
        query_params = []
        if df:
            query_params.append(f"date_from={df}")
        if dt:
            query_params.append(f"date_to={dt}")
        
        query_string = "&".join(query_params)
        
        response_data = {
            "data": rows,
            "download_links": {
                "xlsx": f"{base_url}?file_type=xlsx" + (f"&{query_string}" if query_string else ""),
                "csv": f"{base_url}?file_type=csv" + (f"&{query_string}" if query_string else "")
            }
        }
        
        return Response(response_data)


@extend_schema(
    parameters=[],
    description="AR Aging report. Supports ?as_of=YYYY-MM-DD, and bucket sizes b1,b2,b3.",
    responses={200: None}
)
class ARAgingReport(APIView):
    def get(self, request):
        as_of_str = request.GET.get("as_of")
        if as_of_str:
            try:
                as_of = datetime.strptime(as_of_str, "%Y-%m-%d").date()
            except ValueError:
                as_of = date.today()
        else:
            as_of = date.today()

        b1 = int(request.GET.get("b1", 30))
        b2 = int(request.GET.get("b2", 30))
        b3 = int(request.GET.get("b3", 30))
        fmt = (request.GET.get("format") or "json").lower()
        file_type = request.GET.get("file_type", "").lower()

        data = build_ar_aging(as_of, b1, b2, b3)
        
        if fmt == "csv" or file_type == "csv":
            resp = HttpResponse(content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="ar_aging.csv"'
            resp.write('\ufeff')  # Excel-friendly BOM
            w = csv.writer(resp)
            w.writerow(["Invoice ID","Number","Customer","Date","Due Date","Days Overdue","Bucket","Balance"])
            for r in data["invoices"]:
                w.writerow([r["invoice_id"], r["number"], r["customer"], r["date"], r["due_date"], r["days_overdue"], r["bucket"], f'{r["balance"]:.2f}'])
            # add summary rows
            w.writerow([]); w.writerow(["Summary"])
            for k in data["buckets"] + [">=TOTAL_MARKER"]:
                if k == ">=TOTAL_MARKER":
                    w.writerow(["TOTAL", "", "", "", "", "", "", f'{data["summary"]["TOTAL"]:.2f}'])
                else:
                    w.writerow([k, "", "", "", "", "", "", f'{data["summary"][k]:.2f}'])
            return resp

        if fmt in ("xlsx", "excel") or file_type == "xlsx":
            if not OPENPYXL_OK:
               return HttpResponse("openpyxl not installed", status=400)
            wb = Workbook(); ws = wb.active; ws.title = "AR Aging"
            ws.append(["Invoice ID","Number","Customer","Date","Due Date","Days Overdue","Bucket","Balance"])
            for r in data["invoices"]:
                ws.append([r["invoice_id"], r["number"], r["customer"], r["date"], r["due_date"], r["days_overdue"], r["bucket"], r["balance"]])
            ws.append([]); ws.append(["Summary"])
            for k in data["buckets"]:
                ws.append([k, "", "", "", "", "", "", data["summary"][k]])
            ws.append(["TOTAL","","","","","","", data["summary"]["TOTAL"]])
            bio = BytesIO(); wb.save(bio); bio.seek(0)
            resp = HttpResponse(bio.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = 'attachment; filename="ar_aging.xlsx"'
            return resp

        # Default JSON response with download links
        base_url = request.build_absolute_uri(request.path)
        query_params = []
        if as_of_str:
            query_params.append(f"as_of={as_of_str}")
        if request.GET.get("b1"):
            query_params.append(f"b1={b1}")
        if request.GET.get("b2"):
            query_params.append(f"b2={b2}")
        if request.GET.get("b3"):
            query_params.append(f"b3={b3}")
        
        query_string = "&".join(query_params)
        
        response_data = {
            "data": data,
            "download_links": {
                "xlsx": f"{base_url}?file_type=xlsx" + (f"&{query_string}" if query_string else ""),
                "csv": f"{base_url}?file_type=csv" + (f"&{query_string}" if query_string else "")
            }
        }
        
        return Response(response_data)

@extend_schema(
    parameters=[],
    description="AP Aging report. Supports ?as_of=YYYY-MM-DD, and bucket sizes b1,b2,b3.",
    responses={200: None}
)
class APAgingReport(APIView):
    def get(self, request):
        as_of_str = request.GET.get("as_of")
        if as_of_str:
            as_of_str = as_of_str.strip()  # ✅ remove stray newline
            try:
                as_of = datetime.strptime(as_of_str, "%Y-%m-%d").date()
            except ValueError:
                as_of = date.today()
        else:
            as_of = date.today()

        b1 = int(request.GET.get("b1", 30))
        b2 = int(request.GET.get("b2", 30))
        b3 = int(request.GET.get("b3", 30))
        fmt = (request.GET.get("format") or "json").lower()
        file_type = request.GET.get("file_type", "").lower()

        data = build_ap_aging(as_of, b1, b2, b3)

        if fmt == "csv" or file_type == "csv":
            resp = HttpResponse(content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="ap_aging.csv"'
            resp.write('\ufeff')  # Excel-friendly BOM
            w = csv.writer(resp)
            w.writerow(["Invoice ID","Number","Supplier","Date","Due Date","Days Overdue","Bucket","Balance"])
            for r in data["invoices"]:
                w.writerow([r["id"], r["number"], r["supplier"], r["date"], r["due_date"], r["days_overdue"], r["bucket"], r["balance"]])
            # add summary rows
            w.writerow([]); w.writerow(["Summary"])
            for k in data["buckets"] + [">=TOTAL_MARKER"]:
                if k == ">=TOTAL_MARKER":
                    w.writerow(["TOTAL", "", "", "", "", "", "", f'{data["summary"]["TOTAL"]:.2f}'])
                else:
                    w.writerow([k, "", "", "", "", "", "", f'{data["summary"][k]:.2f}'])
            return resp

        if fmt in ("xlsx", "excel") or file_type == "xlsx":
            if not OPENPYXL_OK:
               return HttpResponse("openpyxl not installed", status=400)
            wb = Workbook(); ws = wb.active; ws.title = "AP Aging"
            ws.append(["Invoice ID","Number","Supplier","Date","Due Date","Days Overdue","Bucket","Balance"])
            for r in data["invoices"]:
                ws.append([r["id"], r["number"], r["supplier"], r["date"], r["due_date"], r["days_overdue"], r["bucket"], r["balance"]])
            ws.append([]); ws.append(["Summary"])
            for k in data["buckets"]:
                ws.append([k, "", "", "", "", "", "", data["summary"][k]])
            ws.append(["TOTAL","","","","","","", data["summary"]["TOTAL"]])
            bio = BytesIO(); wb.save(bio); bio.seek(0)
            resp = HttpResponse(bio.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = 'attachment; filename="ap_aging.xlsx"'
            return resp

        # Default JSON response with download links
        base_url = request.build_absolute_uri(request.path)
        query_params = []
        if as_of_str:
            query_params.append(f"as_of={as_of_str}")
        if request.GET.get("b1"):
            query_params.append(f"b1={b1}")
        if request.GET.get("b2"):
            query_params.append(f"b2={b2}")
        if request.GET.get("b3"):
            query_params.append(f"b3={b3}")
        
        query_string = "&".join(query_params)
        
        response_data = {
            "data": data,
            "download_links": {
                "xlsx": f"{base_url}?file_type=xlsx" + (f"&{query_string}" if query_string else ""),
                "csv": f"{base_url}?file_type=csv" + (f"&{query_string}" if query_string else "")
            }
        }
        
        return Response(response_data)
class BankAccountViewSet(viewsets.ModelViewSet):
    serializer_class = BankAccountSerializer
    queryset = BankAccount.objects.all()


class AccountViewSet(viewsets.ModelViewSet):
    serializer_class = AccountSerializer; queryset = Account.objects.all(); filterset_fields = ["type", "code", "name"]

class JournalEntryViewSet(viewsets.ModelViewSet):
    serializer_class = JournalEntrySerializer; queryset = JournalEntry.objects.all()
    
    @action(detail=True, methods=["post"], url_path="post")
    def post_gl(self, request, pk=None):
        entry = self.get_object(); post_entry(entry); return Response({"status": "posted"})

    @extend_schema(
    parameters=[
        # Beware: drf-spectacular infers types from serializers; use explicit params for simple query strings
        OpenApiExample("CSV export", value={"file_type": "csv"}, parameter_only=True),
        OpenApiExample("XLSX export", value={"file_type": "xlsx"}, parameter_only=True),
    ],
    responses={200: None},
    description="Export journal entries to CSV or XLSX. Query params: file_type=csv|xlsx, date_from, date_to, posted_only=true|false"
)
    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """Export journal entries to CSV or XLSX"""
        file_type = request.GET.get("file_type", "csv").lower()
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        posted_only = request.GET.get("posted_only", "false").lower() == "true"

        # Build queryset with filters
        queryset = JournalEntry.objects.all()
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        if posted_only:
            queryset = queryset.filter(posted=True)
        
        queryset = queryset.order_by('date', 'id').prefetch_related('lines__account')

        if file_type == "csv":
            resp = HttpResponse(content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="journals.csv"'
            resp.write('\ufeff')  # BOM for Excel
            w = csv.writer(resp)
            w.writerow(["Journal ID", "Date", "Currency", "Memo", "Posted", "Account Code", "Account Name", "Debit", "Credit"])
            
            for entry in queryset:
                for line in entry.lines.all():
                    w.writerow([
                        entry.id,
                        entry.date,
                        entry.currency.code if entry.currency else "",
                        entry.memo,
                        "Yes" if entry.posted else "No",
                        line.account.code,
                        line.account.name,
                        f"{line.debit:.2f}",
                        f"{line.credit:.2f}"
                    ])
            return resp

        elif file_type == "xlsx":
            if not OPENPYXL_OK:
                return HttpResponse("openpyxl not installed", status=400)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Journal Entries"
            ws.append(["Journal ID", "Date", "Currency", "Memo", "Posted", "Account Code", "Account Name", "Debit", "Credit"])
            
            for entry in queryset:
                for line in entry.lines.all():
                    ws.append([
                        entry.id,
                        entry.date,
                        entry.currency.code if entry.currency else "",
                        entry.memo,
                        "Yes" if entry.posted else "No",
                        line.account.code,
                        line.account.name,
                        float(line.debit),
                        float(line.credit)
                    ])
            
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            resp = HttpResponse(bio.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = 'attachment; filename="journals.xlsx"'
            return resp
        
        else:
            return Response({"error": "Invalid file_type. Use 'csv' or 'xlsx'."}, status=400)

class ARInvoiceViewSet(viewsets.ModelViewSet):
    serializer_class = ARInvoiceSerializer; queryset = ARInvoice.objects.all()
    @extend_schema(
    responses=inline_serializer(
        name="PostGLResponse",
        fields={
            "already_posted": drf_serializers.BooleanField(),
            "journal": inline_serializer(
                name="JournalEntryInline",
                fields={
                    "id": drf_serializers.IntegerField(),
                    "date": drf_serializers.DateField(),
                    "currency": inline_serializer(
                        name="CurrencyInline",
                        fields={"code": drf_serializers.CharField()}
                    ),
                    "memo": drf_serializers.CharField(allow_null=True, required=False),
                    "posted": drf_serializers.BooleanField(),
                    "lines": drf_serializers.ListField(
                        child=inline_serializer(
                            name="JournalLineInline",
                            fields={
                                "account_code": drf_serializers.CharField(),
                                "account_name": drf_serializers.CharField(),
                                "debit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                                "credit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                            }
                        )
                    ),
                },
            ),
        },
        ref_name="PostGLResponse",
    ),
    examples=[
        OpenApiExample(
            "First time posting — created",
            value={
                "already_posted": False,
                "journal": {
                    "id": 123,
                    "date": "2025-10-01",
                    "currency": {"code": "AED"},
                    "memo": "AR Invoice #INV-001",
                    "posted": True,
                    "lines": [
                        {"account_code": "1100", "account_name": "Accounts Receivable", "debit": 0.00, "credit": 105.00},
                        {"account_code": "4000", "account_name": "Revenue", "debit": 0.00, "credit": 100.00},
                        {"account_code": "2100", "account_name": "VAT Output", "debit": 0.00, "credit": 5.00},
                    ],
                },
            },
            response_only=True,
        ),
        OpenApiExample(
            "Subsequent call — already posted",
            value={
                "already_posted": True,
                "journal": {
                    "id": 123,
                    "date": "2025-10-01",
                    "currency": {"code": "AED"},
                    "memo": "AR Invoice #INV-001",
                    "posted": True,
                    "lines": [
                        {"account_code": "1100", "account_name": "Accounts Receivable", "debit": 0.00, "credit": 105.00},
                        {"account_code": "4000", "account_name": "Revenue", "debit": 0.00, "credit": 100.00},
                        {"account_code": "2100", "account_name": "VAT Output", "debit": 0.00, "credit": 5.00},
                    ],
                },
            },
            response_only=True,
        ),
    ],
)
    @action(detail=True, methods=["post"], url_path="post-gl")
    def post_gl(self, request, pk=None):
        invoice = self.get_object()
        try:
            je, created = gl_post_from_ar_balanced(invoice)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        payload = {"already_posted": not created, "journal": JournalEntryReadSerializer(je).data}
        return Response(payload, status=200 if created or not created else 201)



class APInvoiceViewSet(viewsets.ModelViewSet):
    serializer_class = APInvoiceSerializer; queryset = APInvoice.objects.all()
    @extend_schema(
    responses=inline_serializer(
        name="PostGLResponse",
        fields={
            "already_posted": drf_serializers.BooleanField(),
            "journal": inline_serializer(
                name="JournalEntryInline",
                fields={
                    "id": drf_serializers.IntegerField(),
                    "date": drf_serializers.DateField(),
                    "currency": inline_serializer(
                        name="CurrencyInline",
                        fields={"code": drf_serializers.CharField()}
                    ),
                    "memo": drf_serializers.CharField(allow_null=True, required=False),
                    "posted": drf_serializers.BooleanField(),
                    "lines": drf_serializers.ListField(
                        child=inline_serializer(
                            name="JournalLineInline",
                            fields={
                                "account_code": drf_serializers.CharField(),
                                "account_name": drf_serializers.CharField(),
                                "debit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                                "credit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                            }
                        )
                    ),
                },
            ),
        },
        ref_name="PostGLResponse",
    ),
    examples=[
        OpenApiExample(
            "First time posting — created",
            value={
                "already_posted": False,
                "journal": {
                    "id": 123,
                    "date": "2025-10-01",
                    "currency": {"code": "AED"},
                    "memo": "AR Invoice #INV-001",
                    "posted": True,
                    "lines": [
                        {"account_code": "1100", "account_name": "Accounts Receivable", "debit": 0.00, "credit": 105.00},
                        {"account_code": "4000", "account_name": "Revenue", "debit": 0.00, "credit": 100.00},
                        {"account_code": "2100", "account_name": "VAT Output", "debit": 0.00, "credit": 5.00},
                    ],
                },
            },
            response_only=True,
        ),
        OpenApiExample(
            "Subsequent call — already posted",
            value={
                "already_posted": True,
                "journal": {
                    "id": 123,
                    "date": "2025-10-01",
                    "currency": {"code": "AED"},
                    "memo": "AR Invoice #INV-001",
                    "posted": True,
                    "lines": [
                        {"account_code": "1100", "account_name": "Accounts Receivable", "debit": 0.00, "credit": 105.00},
                        {"account_code": "4000", "account_name": "Revenue", "debit": 0.00, "credit": 100.00},
                        {"account_code": "2100", "account_name": "VAT Output", "debit": 0.00, "credit": 5.00},
                    ],
                },
            },
            response_only=True,
        ),
    ],
)
    @action(detail=True, methods=["post"], url_path="post-gl")
    def post_gl(self, request, pk=None):
        invoice = self.get_object()
        try:
            je, created = gl_post_from_ap_balanced(invoice)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        payload = {"already_posted": not created, "journal": JournalEntryReadSerializer(je).data}
        return Response(payload, status=200 if created or not created else 201)
class ARPaymentViewSet(viewsets.ModelViewSet):
    serializer_class = ARPaymentSerializer
    queryset = ARPayment.objects.all()

    def perform_create(self, serializer):
        payment = serializer.save()  # creates ARPayment
        je, created, closed = post_ar_payment(payment)
        # We return enriched response
        self._last_payload = {
            "journal": JournalEntryReadSerializer(je).data,
            "already_posted": not created,
            "invoice_closed": closed,
        }

    def create(self, request, *args, **kwargs):
        resp = super().create(request, *args, **kwargs)
        # merge with normal serializer output if you like; simplest: replace with posting payload
        return Response(self._last_payload, status=201)

    def perform_update(self, serializer):
        """Reverse old GL entry and create new one when payment is updated"""
        payment = self.get_object()
        old_gl_journal = payment.gl_journal
        old_amount = payment.amount
        
        # Save the updated payment
        payment = serializer.save()
        
        # If GL journal exists and amount changed, reverse and repost
        if old_gl_journal and old_amount != payment.amount:
            reverse_journal(old_gl_journal)
            je, created, closed = post_ar_payment(payment)
            self._last_payload = {
                "journal": JournalEntryReadSerializer(je).data,
                "reversed_journal": old_gl_journal.id,
                "invoice_closed": closed,
            }
        else:
            self._last_payload = {"status": "updated", "no_gl_change": True}

    def update(self, request, *args, **kwargs):
        resp = super().update(request, *args, **kwargs)
        return Response(self._last_payload, status=200)

    def perform_destroy(self, instance):
        """Reverse GL entry when payment is deleted"""
        if instance.gl_journal:
            reverse_journal(instance.gl_journal)
        instance.delete()

    @action(detail=True, methods=["post"])
    def reconcile(self, request, pk=None):
        payment = self.get_object()
        payment.reconciled = True
        payment.reconciliation_ref = request.data.get("reconciliation_ref", "")
        payment.reconciled_at = timezone.now().date()
        payment.save()
        return Response({"status": "reconciled"}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    def unreconcile(self, request, pk=None):
        payment = self.get_object()
        payment.reconciled = False
        payment.reconciliation_ref = ""
        payment.reconciled_at = None
        payment.save()
        return Response({"status": "unreconciled"}, status=status.HTTP_200_OK)
@extend_schema_view(
    create=extend_schema(
        examples=[
            OpenApiExample(
                "Simple AP payment",
                value={"invoice": 456, "date": "2025-10-02", "amount": 115.00, "bank_account": 1},
                request_only=True,
            )
        ],
        responses=inline_serializer(
            name="PaymentResponse",
            fields={
                "already_posted": drf_serializers.BooleanField(),
                "invoice_closed": drf_serializers.BooleanField(),
                "journal": inline_serializer(
                    name="JournalEntryInline",
                    fields={
                        "id": drf_serializers.IntegerField(),
                        "date": drf_serializers.DateField(),
                        "currency": inline_serializer(
                            name="CurrencyInline",
                            fields={"code": drf_serializers.CharField()}
                        ),
                        "memo": drf_serializers.CharField(allow_null=True, required=False),
                        "posted": drf_serializers.BooleanField(),
                        "lines": drf_serializers.ListField(
                            child=inline_serializer(
                                name="JournalLineInline",
                                fields={
                                    "account_code": drf_serializers.CharField(),
                                    "account_name": drf_serializers.CharField(),
                                    "debit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                                    "credit": drf_serializers.DecimalField(max_digits=18, decimal_places=2),
                                }
                            )
                        ),
                    },
                ),
            },
            ref_name="PaymentResponse",
        ),
    )
)
class APPaymentViewSet(viewsets.ModelViewSet):
    serializer_class = APPaymentSerializer
    queryset = APPayment.objects.all()

    def perform_create(self, serializer):
        payment = serializer.save()
        je, created, closed = post_ap_payment(payment)
        self._last_payload = {
            "journal": JournalEntryReadSerializer(je).data,
            "already_posted": not created,
            "invoice_closed": closed,
        }
    @action(detail=True, methods=["post"])
    def reconcile(self, request, pk=None):
        payment = self.get_object()
        payment.reconciled = True
        payment.reconciliation_ref = request.data.get("reconciliation_ref", "")
        payment.reconciled_at = timezone.now().date()
        payment.save()
        return Response({"status": "reconciled"}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    def unreconcile(self, request, pk=None):
        payment = self.get_object()
        payment.reconciled = False
        payment.reconciliation_ref = ""
        payment.reconciled_at = None
        payment.save()
        return Response({"status": "unreconciled"}, status=status.HTTP_200_OK)
    def create(self, request, *args, **kwargs):
        resp = super().create(request, *args, **kwargs)
        return Response(self._last_payload, status=201)

    def perform_update(self, serializer):
        """Reverse old GL entry and create new one when payment is updated"""
        payment = self.get_object()
        old_gl_journal = payment.gl_journal
        old_amount = payment.amount
        
        # Save the updated payment
        payment = serializer.save()
        
        # If GL journal exists and amount changed, reverse and repost
        if old_gl_journal and old_amount != payment.amount:
            reverse_journal(old_gl_journal)
            je, created, closed = post_ap_payment(payment)
            self._last_payload = {
                "journal": JournalEntryReadSerializer(je).data,
                "reversed_journal": old_gl_journal.id,
                "invoice_closed": closed,
            }
        else:
            self._last_payload = {"status": "updated", "no_gl_change": True}

    def update(self, request, *args, **kwargs):
        resp = super().update(request, *args, **kwargs)
        return Response(self._last_payload, status=200)

    def perform_destroy(self, instance):
        """Reverse GL entry when payment is deleted"""
        if instance.gl_journal:
            reverse_journal(instance.gl_journal)
        instance.delete()

# ---- Seed VAT presets ----
class SeedVATPresets(APIView):
    permission_classes = [IsAdminUser]
    def post(self, request):
        s = SeedVATRequestSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        ids = seed_vat_presets(effective_from=s.validated_data.get("effective_from"))
        return Response({"created_ids": ids}, status=201 if ids else 200)

# ---- List tax rates by country ----
class ListTaxRates(APIView):
    def get(self, request):
        country = request.GET.get("country")
        qs = TaxRate.objects.all()
        if country:
            qs = qs.filter(country=country)
        data = [{
            "id": t.id, "name": t.name, "rate": float(t.rate), "country": t.country,
            "category": t.category, "code": t.code, "effective_from": t.effective_from.isoformat() if t.effective_from else None
        } for t in qs.order_by("country","category","rate")]
        return Response(data)

# ---- Corporate tax accrual ----
class CorporateTaxAccrual(APIView):
   # permission_classes = [IsAdminUser]  # Disabled for testing
   def post(self, request):
       s = CorpTaxAccrualRequestSerializer(data=request.data)
       s.is_valid(raise_exception=True)
       country = s.validated_data["country"]
       df = s.validated_data["date_from"]; dt = s.validated_data["date_to"]
       org_id = request.data.get("org_id")
       allow_override = bool(request.data.get("override", False))
       filing, je, meta = accrue_corporate_tax_with_filing(country, df, dt, org_id, allow_override=allow_override)
       if je is None:
           return Response({"created": False, "meta": meta}, status=200)
       return Response({
           "created": True, "filing_id": filing.id,
           "journal": JournalEntryReadSerializer(je).data, "meta": meta
       }, status=201)
   
class CorporateTaxFile(APIView):
   # permission_classes = [IsAdminUser]  # Disabled for testing
   def post(self, request, filing_id:int):
       f = file_corporate_tax(filing_id)
       return Response({"filing_id": f.id, "status": f.status, "filed_at": f.filed_at}, status=200)
class CorporateTaxReverse(APIView):
   # permission_classes = [IsAdminUser]  # Disabled for testing
   def post(self, request, filing_id:int):
       rev = reverse_corporate_tax_filing(filing_id)
       return Response({"reversal_journal": JournalEntryReadSerializer(rev).data}, status=201)
   

class CorporateTaxBreakdown(APIView):
   def get(self, request):
       country = request.GET.get("country")
       df = request.GET.get("date_from"); dt = request.GET.get("date_to")
       org_id = request.GET.get("org_id")
       fmt = (request.GET.get("format") or "json").lower()
       file_type = request.GET.get("file_type", "").lower()
       # recompute using same logic as accrual (but don’t post)
       from datetime import datetime
       df_d = datetime.fromisoformat(df).date() if df else None
       dt_d = datetime.fromisoformat(dt).date() if dt else None
       # collect rows
       qs = JournalLine.objects.select_related("entry","account") \
            .filter(entry__posted=True)
       if df_d: qs = qs.filter(entry__date__gte=df_d)
       if dt_d: qs = qs.filter(entry__date__lte=dt_d)
       qs = _with_org_filter(qs, org_id)
       rows = []
       income = Decimal("0"); expense = Decimal("0")
       for ln in qs:
           kind = "OTHER"
           delta = Decimal("0")
           # Handle both short codes (IN, EX) and full names (INCOME, EXPENSE)
           acc_type = ln.account.type.upper()
           if acc_type in ("INCOME", "IN"):
               kind = "INCOME"
               delta = (Decimal(ln.credit) - Decimal(ln.debit))
               income += delta
           elif acc_type in ("EXPENSE", "EX"):
               kind = "EXPENSE"
               delta = (Decimal(ln.debit) - Decimal(ln.credit))
               expense += delta
           rows.append({
               "date": ln.entry.date.isoformat() if ln.entry.date else None,
               "journal_id": ln.entry.id,
               "account_code": ln.account.code,
               "account_name": ln.account.name,
               "type": kind,
               "delta": float(q2(delta)),
               "debit": float(q2(ln.debit)),
               "credit": float(q2(ln.credit)),
           })
       profit = q2(income - expense)
       meta = {
           "country": country,
           "date_from": df, "date_to": dt, "org_id": org_id,
           "income": float(q2(income)), "expense": float(q2(expense)),
           "profit": float(profit)
       }
       # CSV Export
       if fmt == "csv" or file_type == "csv":
           from django.http import HttpResponse; import csv
           resp = HttpResponse(content_type="text/csv; charset=utf-8")
           resp["Content-Disposition"] = 'attachment; filename="corp_tax_breakdown.csv"'
           resp.write('\ufeff')  # Excel-friendly BOM
           w = csv.writer(resp)
           w.writerow(["Date","Journal","Account Code","Account Name","Type","Delta","Debit","Credit"])
           for r in rows: w.writerow([r["date"], r["journal_id"], r["account_code"], r["account_name"], r["type"], f'{r["delta"]:.2f}', f'{r["debit"]:.2f}', f'{r["credit"]:.2f}'])
           # footer
           w.writerow([]); w.writerow(["Income", f'{meta["income"]:.2f}'])
           w.writerow(["Expense", f'{meta["expense"]:.2f}'])
           w.writerow(["Profit", f'{meta["profit"]:.2f}'])
           return resp
       # XLSX Export
       if fmt in ("xlsx","excel") or file_type == "xlsx":
           if not OPENPYXL_OK:
               return HttpResponse("openpyxl not installed", status=400)
           from io import BytesIO
           wb = Workbook(); ws = wb.active; ws.title="CorpTax Breakdown"
           ws.append(["Date","Journal","Account Code","Account Name","Type","Delta","Debit","Credit"])
           for r in rows: ws.append([r["date"], r["journal_id"], r["account_code"], r["account_name"], r["type"], r["delta"], r["debit"], r["credit"]])
           ws.append([]); ws.append(["Income", meta["income"]]); ws.append(["Expense", meta["expense"]]); ws.append(["Profit", meta["profit"]])
           bio = BytesIO(); wb.save(bio); bio.seek(0)
           from django.http import HttpResponse
           resp = HttpResponse(bio.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
           resp["Content-Disposition"] = 'attachment; filename="corp_tax_breakdown.xlsx"'
           return resp
       
       # Default JSON response with download links (like Trail Balance)
       base_url = request.build_absolute_uri(request.path)
       query_params = []
       if country:
           query_params.append(f"country={country}")
       if df:
           query_params.append(f"date_from={df}")
       if dt:
           query_params.append(f"date_to={dt}")
       if org_id:
           query_params.append(f"org_id={org_id}")
       
       query_string = "&".join(query_params)
       
       response_data = {
           "meta": meta,
           "rows": rows,
           "download_links": {
               "xlsx": f"{base_url}?file_type=xlsx" + (f"&{query_string}" if query_string else ""),
               "csv": f"{base_url}?file_type=csv" + (f"&{query_string}" if query_string else "")
           }
       }
       
       return Response(response_data, status=200)

class InvoiceViewSet(viewsets.ModelViewSet):
    queryset = Invoice.objects.prefetch_related("lines")
    serializer_class = InvoiceSerializer

    @action(detail=True, methods=["POST"])
    def post(self, request, pk=None):
        inv = post_invoice(pk)
        return Response(InvoiceSerializer(inv).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["POST"])
    def reverse(self, request, pk=None):
        rev = reverse_posted_invoice(pk)
        return Response(InvoiceSerializer(rev).data, status=status.HTTP_201_CREATED)